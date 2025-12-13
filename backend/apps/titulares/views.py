from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from django_filters.rest_framework import DjangoFilterBackend
from django_filters import rest_framework as django_filters
from django.db import transaction
from django.db.models import Q
import openpyxl
from io import BytesIO
from .models import Titular, VinculoTitular, Dependente, VinculoDependente
from .serializers import (
    TitularSerializer, TitularListSerializer, TitularCreateUpdateSerializer,
    VinculoTitularSerializer, DependenteSerializer, VinculoDependenteSerializer
)
from apps.core.models import AmparoLegal
from apps.accounts.permissions import CargoBasedPermission, PermissionMessageMixin, IsGestorOuSuperior


class TitularFilter(django_filters.FilterSet):
    """Filtro customizado para Titular."""
    empresa = django_filters.UUIDFilter(method='filter_by_empresa')
    tipo_vinculo = django_filters.CharFilter(method='filter_by_tipo_vinculo')
    vinculo_status = django_filters.BooleanFilter(method='filter_by_vinculo_status')
    data_fim_vinculo_gte = django_filters.DateFilter(method='filter_data_fim_gte')
    data_fim_vinculo_lte = django_filters.DateFilter(method='filter_data_fim_lte')
    data_entrada_gte = django_filters.DateFilter(method='filter_data_entrada_gte')
    data_entrada_lte = django_filters.DateFilter(method='filter_data_entrada_lte')
    ultima_atualizacao_gte = django_filters.DateFilter(method='filter_ultima_atualizacao_gte')
    ultima_atualizacao_lte = django_filters.DateFilter(method='filter_ultima_atualizacao_lte')
    
    class Meta:
        model = Titular
        fields = ['nacionalidade', 'sexo']
    
    def filter_by_empresa(self, queryset, name, value):
        """Filtra titulares que têm vínculo com a empresa especificada."""
        if value:
            return queryset.filter(vinculos__empresa=value).distinct()
        return queryset
    
    def filter_by_tipo_vinculo(self, queryset, name, value):
        """Filtra titulares pelo tipo de vínculo."""
        if value:
            return queryset.filter(vinculos__tipo_vinculo=value).distinct()
        return queryset
    
    def filter_by_vinculo_status(self, queryset, name, value):
        """Filtra titulares pelo status do vínculo."""
        if value is not None:
            return queryset.filter(vinculos__status=value).distinct()
        return queryset
    
    def filter_data_fim_gte(self, queryset, name, value):
        """Filtra titulares com data fim vínculo >= valor."""
        if value:
            return queryset.filter(vinculos__data_fim_vinculo__gte=value).distinct()
        return queryset
    
    def filter_data_fim_lte(self, queryset, name, value):
        """Filtra titulares com data fim vínculo <= valor."""
        if value:
            return queryset.filter(vinculos__data_fim_vinculo__lte=value).distinct()
        return queryset
    
    def filter_data_entrada_gte(self, queryset, name, value):
        """Filtra titulares com data entrada no país >= valor."""
        if value:
            return queryset.filter(vinculos__data_entrada_pais__gte=value).distinct()
        return queryset
    
    def filter_data_entrada_lte(self, queryset, name, value):
        """Filtra titulares com data entrada no país <= valor."""
        if value:
            return queryset.filter(vinculos__data_entrada_pais__lte=value).distinct()
        return queryset
    
    def filter_ultima_atualizacao_gte(self, queryset, name, value):
        """Filtra titulares com última atualização >= valor."""
        if value:
            return queryset.filter(vinculos__ultima_atualizacao__gte=value).distinct()
        return queryset
    
    def filter_ultima_atualizacao_lte(self, queryset, name, value):
        """Filtra titulares com última atualização <= valor."""
        if value:
            return queryset.filter(vinculos__ultima_atualizacao__lte=value).distinct()
        return queryset


class TitularViewSet(PermissionMessageMixin, viewsets.ModelViewSet):
    """
    ViewSet para gerenciamento de titulares.
    
    Permissões:
    - Consultor: Apenas visualização (list, retrieve)
    - Gestor: CRUD completo (create, update, delete)
    - Diretor: CRUD + admin
    """
    
    queryset = Titular.objects.select_related(
        'criado_por', 'atualizado_por'
    ).prefetch_related('vinculos', 'vinculos__empresa', 'vinculos__amparo', 'dependentes')
    permission_classes = [IsAuthenticated, CargoBasedPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = TitularFilter
    search_fields = ['nome', 'rnm', 'cpf', 'passaporte', 'email']
    ordering_fields = ['nome', 'rnm', 'data_criacao', 'data_nascimento']
    ordering = ['nome']
    
    def get_serializer_class(self):
        if self.action == 'list':
            return TitularListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return TitularCreateUpdateSerializer
        return TitularSerializer
    
    def perform_create(self, serializer):
        serializer.save(criado_por=self.request.user, atualizado_por=self.request.user)
    
    def perform_update(self, serializer):
        serializer.save(atualizado_por=self.request.user)
    
    @action(detail=True, methods=['get'])
    def vinculos(self, request, pk=None):
        """Retorna todos os vínculos de um titular."""
        titular = self.get_object()
        vinculos = titular.vinculos.all()
        serializer = VinculoTitularSerializer(vinculos, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def dependentes(self, request, pk=None):
        """Retorna todos os dependentes de um titular."""
        titular = self.get_object()
        dependentes = titular.dependentes.all()
        serializer = DependenteSerializer(dependentes, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser],
            permission_classes=[IsAuthenticated, IsGestorOuSuperior])
    def importar(self, request):
        """
        Importa titulares de uma planilha Excel.
        Requer permissão de Gestor ou superior.
        """
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'Arquivo não enviado'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            wb = openpyxl.load_workbook(BytesIO(file.read()))
            ws = wb.active
            
            # Pegar cabeçalhos da primeira linha
            headers = [cell.value.strip().lower() if cell.value else '' for cell in ws[1]]
            
            # Mapear colunas
            col_map = {}
            header_map = {
                'nome': 'nome',
                'nacionalidade': 'nacionalidade',
                'nascimento': 'data_nascimento',
                'mãe': 'mae',
                'mae': 'mae',
                'pai': 'pai',
                'rnm': 'rnm',
                'amparo': 'amparo',
                'prazo': 'data_fim_vinculo',
                'status': 'status',
            }
            
            for idx, header in enumerate(headers):
                if header in header_map:
                    col_map[header_map[header]] = idx
            
            # Cache de amparos
            amparos = {a.nome.lower(): a for a in AmparoLegal.objects.all()}
            
            titulares_criados = 0
            titulares_atualizados = 0
            erros = []
            
            with transaction.atomic():
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        # Extrair dados da linha
                        nome = row[col_map.get('nome', 0)] if col_map.get('nome') is not None else None
                        if not nome:
                            continue
                        
                        rnm = row[col_map.get('rnm')] if col_map.get('rnm') is not None else None
                        
                        # Obter nacionalidade como texto
                        nacionalidade = None
                        if col_map.get('nacionalidade') is not None:
                            nac_nome = row[col_map['nacionalidade']]
                            if nac_nome:
                                nacionalidade = str(nac_nome).strip().upper()
                        
                        # Buscar ou criar titular
                        titular = None
                        if rnm:
                            titular = Titular.objects.filter(rnm=rnm).first()
                        
                        if titular:
                            # Atualizar titular existente
                            titular.nome = nome
                            if nacionalidade:
                                titular.nacionalidade = nacionalidade
                            if col_map.get('data_nascimento') is not None:
                                data_nasc = row[col_map['data_nascimento']]
                                if data_nasc:
                                    titular.data_nascimento = data_nasc
                            if col_map.get('pai') is not None:
                                pai = row[col_map['pai']]
                                if pai:
                                    titular.filiacao_um = pai
                            if col_map.get('mae') is not None:
                                mae = row[col_map['mae']]
                                if mae:
                                    titular.filiacao_dois = mae
                            titular.atualizado_por = request.user
                            titular.save()
                            titulares_atualizados += 1
                        else:
                            # Criar novo titular
                            titular_data = {
                                'nome': nome,
                                'rnm': rnm,
                                'nacionalidade': nacionalidade,
                                'criado_por': request.user,
                                'atualizado_por': request.user,
                            }
                            if col_map.get('data_nascimento') is not None:
                                data_nasc = row[col_map['data_nascimento']]
                                if data_nasc:
                                    titular_data['data_nascimento'] = data_nasc
                            if col_map.get('pai') is not None:
                                pai = row[col_map['pai']]
                                if pai:
                                    titular_data['filiacao_um'] = pai
                            if col_map.get('mae') is not None:
                                mae = row[col_map['mae']]
                                if mae:
                                    titular_data['filiacao_dois'] = mae
                            
                            titular = Titular.objects.create(**titular_data)
                            titulares_criados += 1
                        
                        # Criar/Atualizar vínculo se houver amparo e prazo
                        if col_map.get('amparo') is not None and col_map.get('data_fim_vinculo') is not None:
                            amparo_nome = row[col_map['amparo']]
                            data_fim = row[col_map['data_fim_vinculo']]
                            status_val = row[col_map.get('status')] if col_map.get('status') is not None else 'Ativo'
                            
                            if amparo_nome and data_fim:
                                amparo = amparos.get(str(amparo_nome).lower())
                                vinculo_status = str(status_val).lower() in ['ativo', 'true', '1', 'sim'] if status_val else True
                                
                                # Verificar se já existe vínculo com mesmo amparo
                                vinculo_existente = VinculoTitular.objects.filter(
                                    titular=titular,
                                    amparo=amparo
                                ).first() if amparo else None
                                
                                if vinculo_existente:
                                    vinculo_existente.data_fim_vinculo = data_fim
                                    vinculo_existente.status = vinculo_status
                                    vinculo_existente.atualizado_por = request.user
                                    vinculo_existente.save()
                                elif amparo:
                                    VinculoTitular.objects.create(
                                        titular=titular,
                                        amparo=amparo,
                                        tipo_vinculo='PARTICULAR',
                                        data_fim_vinculo=data_fim,
                                        status=vinculo_status,
                                        criado_por=request.user,
                                        atualizado_por=request.user,
                                    )
                    
                    except Exception as e:
                        erros.append(f'Linha {row_num}: {str(e)}')
            
            return Response({
                'message': 'Importação concluída',
                'titulares_criados': titulares_criados,
                'titulares_atualizados': titulares_atualizados,
                'erros': erros[:10] if erros else []  # Limitar erros retornados
            })
            
        except Exception as e:
            return Response({'error': f'Erro ao processar arquivo: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)


class VinculoTitularViewSet(PermissionMessageMixin, viewsets.ModelViewSet):
    """
    ViewSet para gerenciamento de vínculos de titulares.
    
    Permissões baseadas no Cargo do usuário.
    """
    
    queryset = VinculoTitular.objects.select_related(
        'titular', 'empresa', 'amparo', 
        'tipo_atualizacao', 'criado_por', 'atualizado_por'
    )
    serializer_class = VinculoTitularSerializer
    permission_classes = [IsAuthenticated, CargoBasedPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = {
        'titular': ['exact'],
        'tipo_vinculo': ['exact'],
        'status': ['exact'],
        'empresa': ['exact'],
        'data_fim_vinculo': ['exact', 'gte', 'lte', 'gt', 'lt'],
        'data_entrada_pais': ['exact', 'gte', 'lte', 'gt', 'lt'],
    }
    search_fields = ['titular__nome', 'titular__rnm', 'empresa__nome', 'observacoes']
    ordering_fields = ['data_criacao', 'data_entrada_pais', 'data_fim_vinculo', 'titular__nome']
    ordering = ['-data_criacao']
    
    def perform_create(self, serializer):
        serializer.save(criado_por=self.request.user, atualizado_por=self.request.user)
    
    def perform_update(self, serializer):
        serializer.save(atualizado_por=self.request.user)


class DependenteViewSet(PermissionMessageMixin, viewsets.ModelViewSet):
    """
    ViewSet para gerenciamento de dependentes.
    
    Permissões baseadas no Cargo do usuário.
    """
    
    queryset = Dependente.objects.select_related(
        'titular', 'criado_por', 'atualizado_por'
    )
    serializer_class = DependenteSerializer
    permission_classes = [IsAuthenticated, CargoBasedPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['titular', 'tipo_dependente', 'nacionalidade', 'sexo']
    search_fields = ['nome', 'rnm', 'passaporte', 'titular__nome']
    ordering_fields = ['nome', 'data_criacao', 'data_nascimento']
    ordering = ['nome']
    
    def perform_create(self, serializer):
        serializer.save(criado_por=self.request.user, atualizado_por=self.request.user)
    
    def perform_update(self, serializer):
        serializer.save(atualizado_por=self.request.user)


class VinculoDependenteViewSet(PermissionMessageMixin, viewsets.ModelViewSet):
    """
    ViewSet para gerenciamento de vínculos de dependentes.
    
    Permissões baseadas no Cargo do usuário.
    """
    
    queryset = VinculoDependente.objects.select_related(
        'dependente', 'amparo', 'tipo_atualizacao',
        'criado_por', 'atualizado_por'
    )
    serializer_class = VinculoDependenteSerializer
    permission_classes = [IsAuthenticated, CargoBasedPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['dependente', 'status', 'amparo', 'consulado', 'tipo_atualizacao']
    search_fields = ['dependente__nome', 'observacoes']
    ordering_fields = ['data_fim_vinculo', 'data_entrada', 'data_criacao']
    ordering = ['-data_criacao']
    
    def perform_create(self, serializer):
        serializer.save(criado_por=self.request.user, atualizado_por=self.request.user)
    
    def perform_update(self, serializer):
        serializer.save(atualizado_por=self.request.user)


class PesquisaUnificadaViewSet(PermissionMessageMixin, viewsets.ViewSet):
    """
    ViewSet para pesquisa unificada de titulares e dependentes com paginação real.
    Retorna titulares com seus vínculos e dependentes de forma paginada.
    
    Apenas visualização (não permite criar/editar/deletar).
    """
    permission_classes = [IsAuthenticated, CargoBasedPermission]
    
    # Define o modelo para a verificação de permissões
    queryset = Titular.objects.none()
    
    def list(self, request):
        """
        Pesquisa unificada paginada.
        
        Parâmetros:
        - search: termo de busca (nome, rnm, cpf, passaporte)
        - search_field: campo específico para busca (nome, rnm, cpf, passaporte)
        - tipo: filtrar por tipo (titular, dependente)
        - nacionalidade: ID da nacionalidade
        - empresa: ID da empresa
        - tipo_vinculo: EMPRESA ou PARTICULAR
        - vinculo_status: true/false
        - tipo_evento: vencimento, entrada, atualizacao
        - data_de: data início do filtro (YYYY-MM-DD)
        - data_ate: data fim do filtro (YYYY-MM-DD)
        - page: número da página
        - page_size: itens por página (default 20, max 100)
        """
        from django.core.paginator import Paginator
        from django.db.models import Count, Prefetch
        
        # Parâmetros de paginação
        page = int(request.query_params.get('page', 1))
        page_size = min(int(request.query_params.get('page_size', 20)), 100)
        
        # Parâmetros de filtro
        search = request.query_params.get('search', '').strip()
        search_field = request.query_params.get('search_field', '').strip()
        tipo_registro = request.query_params.get('tipo', '').strip().lower()  # titular ou dependente
        nacionalidade = request.query_params.get('nacionalidade')
        consulado = request.query_params.get('consulado')
        empresa = request.query_params.get('empresa')
        tipo_vinculo = request.query_params.get('tipo_vinculo')
        vinculo_status = request.query_params.get('vinculo_status')
        tipo_evento = request.query_params.get('tipo_evento')
        data_de = request.query_params.get('data_de')
        data_ate = request.query_params.get('data_ate')
        
        # Query base para titulares
        titulares_qs = Titular.objects.prefetch_related(
            Prefetch(
                'vinculos',
                queryset=VinculoTitular.objects.select_related('empresa', 'amparo').order_by('-data_fim_vinculo')
            ),
            Prefetch(
                'dependentes',
                queryset=Dependente.objects.prefetch_related(
                    Prefetch(
                        'vinculos',
                        queryset=VinculoDependente.objects.filter(status=True).order_by('-data_fim_vinculo')
                    )
                )
            )
        )
        
        # Query base para dependentes (busca independente)
        dependentes_qs = Dependente.objects.select_related('titular').prefetch_related(
            Prefetch(
                'vinculos',
                queryset=VinculoDependente.objects.filter(status=True).order_by('-data_fim_vinculo')
            )
        )
        
        # Aplicar filtro de busca
        if search:
            # Se tem campo específico, buscar apenas nesse campo
            if search_field == 'nome':
                titulares_qs = titulares_qs.filter(nome__icontains=search)
                dependentes_qs = dependentes_qs.filter(nome__icontains=search)
            elif search_field == 'rnm':
                titulares_qs = titulares_qs.filter(rnm__icontains=search)
                dependentes_qs = dependentes_qs.filter(rnm__icontains=search)
            elif search_field == 'cpf':
                titulares_qs = titulares_qs.filter(cpf__icontains=search)
                dependentes_qs = dependentes_qs.filter(cpf__icontains=search)
            elif search_field == 'passaporte':
                titulares_qs = titulares_qs.filter(passaporte__icontains=search)
                dependentes_qs = dependentes_qs.filter(passaporte__icontains=search)
            else:
                # Busca em todos os campos
                titulares_qs = titulares_qs.filter(
                    Q(nome__icontains=search) |
                    Q(rnm__icontains=search) |
                    Q(cpf__icontains=search) |
                    Q(passaporte__icontains=search)
                )
                dependentes_qs = dependentes_qs.filter(
                    Q(nome__icontains=search) |
                    Q(rnm__icontains=search) |
                    Q(passaporte__icontains=search)
                )
        
        # Filtro de nacionalidade (agora é campo texto)
        if nacionalidade:
            titulares_qs = titulares_qs.filter(nacionalidade__icontains=nacionalidade)
            dependentes_qs = dependentes_qs.filter(nacionalidade__icontains=nacionalidade)
        
        # Filtro de consulado (agora é campo texto no vínculo)
        if consulado:
            titulares_qs = titulares_qs.filter(vinculos__consulado__icontains=consulado).distinct()
            dependentes_qs = dependentes_qs.filter(vinculos__consulado__icontains=consulado).distinct()
        
        # Filtro de empresa (só se aplica a titulares)
        if empresa:
            titulares_qs = titulares_qs.filter(vinculos__empresa_id=empresa).distinct()
        
        # Filtro de tipo de vínculo
        if tipo_vinculo:
            titulares_qs = titulares_qs.filter(vinculos__tipo_vinculo=tipo_vinculo).distinct()
        
        # Filtro de status do vínculo
        if vinculo_status is not None:
            status_bool = vinculo_status.lower() == 'true'
            titulares_qs = titulares_qs.filter(vinculos__status=status_bool).distinct()
        
        # Filtros de data (baseado no tipo de evento)
        if tipo_evento and (data_de or data_ate):
            if tipo_evento == 'vencimento':
                if data_de:
                    titulares_qs = titulares_qs.filter(vinculos__data_fim_vinculo__gte=data_de).distinct()
                if data_ate:
                    titulares_qs = titulares_qs.filter(vinculos__data_fim_vinculo__lte=data_ate).distinct()
            elif tipo_evento == 'entrada':
                if data_de:
                    titulares_qs = titulares_qs.filter(vinculos__data_entrada_pais__gte=data_de).distinct()
                if data_ate:
                    titulares_qs = titulares_qs.filter(vinculos__data_entrada_pais__lte=data_ate).distinct()
            elif tipo_evento == 'atualizacao':
                if data_de:
                    titulares_qs = titulares_qs.filter(vinculos__atualizacao__gte=data_de).distinct()
                if data_ate:
                    titulares_qs = titulares_qs.filter(vinculos__atualizacao__lte=data_ate).distinct()
        
        # Ordenar titulares
        titulares_qs = titulares_qs.order_by('nome')
        
        # Se filtrar apenas dependentes, não buscar titulares
        if tipo_registro == 'dependente':
            titulares_qs = Titular.objects.none()
        
        # Se filtrar apenas titulares, não buscar dependentes
        if tipo_registro == 'titular':
            dependentes_qs = Dependente.objects.none()
        
        # Caso especial: apenas dependentes
        if tipo_registro == 'dependente':
            # Ordenar dependentes
            dependentes_qs = dependentes_qs.order_by('nome')
            total_dependentes = dependentes_qs.count()
            
            # Paginar dependentes
            paginator = Paginator(dependentes_qs, page_size)
            dependentes_page = paginator.get_page(page)
            
            results = []
            for dep in dependentes_page:
                # Pegar o vínculo ativo mais recente do dependente
                dep_vinculos = list(dep.vinculos.all())
                dep_vinculo_ativo = dep_vinculos[0] if dep_vinculos else None
                
                results.append({
                    'type': 'dependente',
                    'id': str(dep.id),
                    'visibleId': f'dependente-{dep.id}',
                    'titularId': str(dep.titular_id) if dep.titular_id else None,
                    'titularNome': dep.titular.nome if dep.titular else 'Sem Titular',
                    'nome': dep.nome,
                    'rnm': dep.rnm,
                    'passaporte': dep.passaporte,
                    'nacionalidade': dep.nacionalidade,
                    'tipoDependente': dep.get_tipo_dependente_display(),
                    'dataNascimento': str(dep.data_nascimento) if dep.data_nascimento else None,
                    'filiacao_um': dep.filiacao_um,
                    'filiacao_dois': dep.filiacao_dois,
                    'dataFimVinculo': str(dep_vinculo_ativo.data_fim_vinculo) if dep_vinculo_ativo and dep_vinculo_ativo.data_fim_vinculo else None,
                    'amparo': dep_vinculo_ativo.amparo.nome if dep_vinculo_ativo and dep_vinculo_ativo.amparo else None,
                })
            
            return Response({
                'results': results,
                'count': total_dependentes,
                'total_records': len(results),
                'page': page,
                'page_size': page_size,
                'total_pages': paginator.num_pages,
                'has_next': dependentes_page.has_next(),
                'has_previous': dependentes_page.has_previous(),
            })
        
        # Contar totais antes de paginar
        total_titulares = titulares_qs.count()
        
        # Paginar titulares
        paginator = Paginator(titulares_qs, page_size)
        titulares_page = paginator.get_page(page)
        
        # Montar resultado
        results = []
        titular_ids_na_pagina = set()
        
        for titular in titulares_page:
            titular_ids_na_pagina.add(titular.id)
            vinculos = list(titular.vinculos.all())
            dependentes = list(titular.dependentes.all())
            
            # Filtrar vínculos baseado nos parâmetros de filtro
            vinculos_filtrados = vinculos
            
            # Filtro de tipo de vínculo
            if tipo_vinculo:
                vinculos_filtrados = [v for v in vinculos_filtrados if v.tipo_vinculo == tipo_vinculo]
            
            # Filtro de consulado (agora é campo texto)
            if consulado:
                vinculos_filtrados = [v for v in vinculos_filtrados if v.consulado and consulado.upper() in v.consulado.upper()]
            
            # Filtro de empresa (empresa_id é UUID)
            if empresa:
                from uuid import UUID
                try:
                    empresa_uuid = UUID(empresa) if isinstance(empresa, str) else empresa
                    vinculos_filtrados = [v for v in vinculos_filtrados if v.empresa_id == empresa_uuid]
                except (ValueError, TypeError):
                    vinculos_filtrados = []
            
            # Filtro de status
            if vinculo_status is not None:
                status_bool = vinculo_status.lower() == 'true'
                vinculos_filtrados = [v for v in vinculos_filtrados if v.status == status_bool]
            
            # Filtro de data (baseado no tipo de evento)
            if tipo_evento and (data_de or data_ate):
                vinculos_temp = []
                for v in vinculos_filtrados:
                    if tipo_evento == 'vencimento':
                        data_campo = v.data_fim_vinculo
                    elif tipo_evento == 'entrada':
                        data_campo = v.data_entrada_pais
                    elif tipo_evento == 'atualizacao':
                        data_campo = v.atualizacao
                    else:
                        data_campo = None
                    
                    if data_campo:
                        from datetime import datetime
                        data_campo_str = data_campo.strftime('%Y-%m-%d') if hasattr(data_campo, 'strftime') else str(data_campo)
                        
                        passa_de = not data_de or data_campo_str >= data_de
                        passa_ate = not data_ate or data_campo_str <= data_ate
                        
                        if passa_de and passa_ate:
                            vinculos_temp.append(v)
                    else:
                        # Se não tem a data do campo filtrado, não incluir
                        pass
                vinculos_filtrados = vinculos_temp
            
            vinculos = vinculos_filtrados
            
            # Verificar se há filtros de vínculo aplicados
            has_vinculo_filters = tipo_vinculo or consulado or empresa or vinculo_status is not None or (tipo_evento and (data_de or data_ate))
            
            if not vinculos:
                # Se há filtros de vínculo ativos e nenhum vínculo passou, não mostrar o titular
                if has_vinculo_filters:
                    continue
                
                # Titular sem vínculo (sem filtros de vínculo ativos)
                results.append({
                    'type': 'titular',
                    'id': str(titular.id),
                    'visibleId': f'titular-{titular.id}-0',
                    'nome': titular.nome,
                    'rnm': titular.rnm,
                    'cpf': titular.cpf,
                    'passaporte': titular.passaporte,
                    'nacionalidade': titular.nacionalidade,
                    'tipoVinculo': None,
                    'empresa': None,
                    'amparo': None,
                    'dataFimVinculo': None,
                    'status': None,
                    'vinculoId': None,
                    'email': titular.email,
                    'telefone': titular.telefone,
                    'filiacao_um': titular.filiacao_um,
                    'filiacao_dois': titular.filiacao_dois,
                    'dataNascimento': str(titular.data_nascimento) if titular.data_nascimento else None,
                    'isLastVinculo': True,
                })
                
                # Adicionar dependentes após titular sem vínculo (apenas se não estiver filtrando só titulares)
                if tipo_registro != 'titular':
                    for dep in dependentes:
                        # Pegar o vínculo ativo mais recente do dependente
                        dep_vinculos = list(dep.vinculos.all())
                        dep_vinculo_ativo = dep_vinculos[0] if dep_vinculos else None
                        
                        results.append({
                            'type': 'dependente',
                            'id': str(dep.id),
                            'visibleId': f'dependente-{dep.id}',
                            'titularId': str(titular.id),
                            'titularNome': titular.nome,
                            'nome': dep.nome,
                            'rnm': dep.rnm,
                            'passaporte': dep.passaporte,
                            'nacionalidade': dep.nacionalidade,
                            'tipoDependente': dep.get_tipo_dependente_display(),
                            'dataNascimento': str(dep.data_nascimento) if dep.data_nascimento else None,
                            'filiacao_um': dep.filiacao_um,
                            'filiacao_dois': dep.filiacao_dois,
                            'dataFimVinculo': str(dep_vinculo_ativo.data_fim_vinculo) if dep_vinculo_ativo and dep_vinculo_ativo.data_fim_vinculo else None,
                            'amparo': dep_vinculo_ativo.amparo.nome if dep_vinculo_ativo and dep_vinculo_ativo.amparo else None,
                        })
            else:
                # Uma linha para cada vínculo
                for idx, vinculo in enumerate(vinculos):
                    is_last = idx == len(vinculos) - 1
                    
                    results.append({
                        'type': 'titular',
                        'id': str(titular.id),
                        'visibleId': f'titular-{titular.id}-{vinculo.id}',
                        'nome': titular.nome,
                        'rnm': titular.rnm,
                        'cpf': titular.cpf,
                        'passaporte': titular.passaporte,
                        'nacionalidade': titular.nacionalidade,
                        'tipoVinculo': vinculo.get_tipo_vinculo_display(),
                        'empresa': vinculo.empresa.nome if vinculo.empresa else None,
                        'amparo': vinculo.amparo.nome if vinculo.amparo else None,
                        'dataFimVinculo': str(vinculo.data_fim_vinculo) if vinculo.data_fim_vinculo else None,
                        'status': vinculo.status,
                        'vinculoId': str(vinculo.id),
                        'email': titular.email,
                        'telefone': titular.telefone,
                        'filiacao_um': titular.filiacao_um,
                        'filiacao_dois': titular.filiacao_dois,
                        'dataNascimento': str(titular.data_nascimento) if titular.data_nascimento else None,
                        'isLastVinculo': is_last,
                    })
                    
                    # Adicionar dependentes após o último vínculo (apenas se não estiver filtrando só titulares)
                    if is_last and tipo_registro != 'titular':
                        for dep in dependentes:
                            # Pegar o vínculo ativo mais recente do dependente
                            dep_vinculos = list(dep.vinculos.all())
                            dep_vinculo_ativo = dep_vinculos[0] if dep_vinculos else None
                            
                            results.append({
                                'type': 'dependente',
                                'id': str(dep.id),
                                'visibleId': f'dependente-{dep.id}',
                                'titularId': str(titular.id),
                                'titularNome': titular.nome,
                                'nome': dep.nome,
                                'rnm': dep.rnm,
                                'passaporte': dep.passaporte,
                                'nacionalidade': dep.nacionalidade,
                                'tipoDependente': dep.get_tipo_dependente_display(),
                                'dataNascimento': str(dep.data_nascimento) if dep.data_nascimento else None,
                                'filiacao_um': dep.filiacao_um,
                                'filiacao_dois': dep.filiacao_dois,
                                'dataFimVinculo': str(dep_vinculo_ativo.data_fim_vinculo) if dep_vinculo_ativo and dep_vinculo_ativo.data_fim_vinculo else None,
                                'amparo': dep_vinculo_ativo.amparo.nome if dep_vinculo_ativo and dep_vinculo_ativo.amparo else None,
                            })
        
        # Se há termo de busca, verificar dependentes órfãos (cujo titular não está na página)
        # Mas não se estiver filtrando apenas titulares
        if search and tipo_registro != 'titular':
            dependentes_orfaos = dependentes_qs.exclude(titular_id__in=titular_ids_na_pagina)
            for dep in dependentes_orfaos[:20]:  # Limitar órfãos por página
                # Pegar o vínculo ativo mais recente do dependente
                dep_vinculos = list(dep.vinculos.all())
                dep_vinculo_ativo = dep_vinculos[0] if dep_vinculos else None
                
                results.append({
                    'type': 'dependente-orphan',
                    'id': str(dep.id),
                    'visibleId': f'dependente-orphan-{dep.id}',
                    'titularId': str(dep.titular_id),
                    'titularNome': dep.titular.nome if dep.titular else 'Desconhecido',
                    'nome': dep.nome,
                    'rnm': dep.rnm,
                    'passaporte': dep.passaporte,
                    'nacionalidade': dep.nacionalidade,
                    'tipoDependente': dep.get_tipo_dependente_display(),
                    'dataNascimento': str(dep.data_nascimento) if dep.data_nascimento else None,
                    'filiacao_um': dep.filiacao_um,
                    'filiacao_dois': dep.filiacao_dois,
                    'dataFimVinculo': str(dep_vinculo_ativo.data_fim_vinculo) if dep_vinculo_ativo and dep_vinculo_ativo.data_fim_vinculo else None,
                    'amparo': dep_vinculo_ativo.amparo.nome if dep_vinculo_ativo and dep_vinculo_ativo.amparo else None,
                })
        
        return Response({
            'results': results,
            'count': total_titulares,
            'total_records': len(results),  # Total de registros na página (titulares + dependentes)
            'page': page,
            'page_size': page_size,
            'total_pages': paginator.num_pages,
            'has_next': titulares_page.has_next(),
            'has_previous': titulares_page.has_previous(),
        })
